import re
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from collections import defaultdict
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import TruncatedSVD
import numpy as np
from nltk.corpus import stopwords

stop_words = set(stopwords.words('english'))
stop_words.update(stopwords.words('turkish'))


# Dosya okuma ve metnin hazırlanması
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Abstracts/Journal of Nursology/Covid/10.17049-ataunihem.749206-1180772.txt", 'r',encoding="utf-8") as text2:
    abstract=text2.read()
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/İşlenecek Kısım/Journal of Nursology/Covid/10.17049-ataunihem.749206-1180772.txt", 'r',encoding="utf-8") as text:
    islenecek=text.read()
    text=islenecek
    #print(islenecek)
    
    
#önişlem aşaması        
#boşukların, parantezlerin silinmesi    


abstract=abstract.lower()
abstract = re.sub(r'\([^()]*\d+[^()]*\)', '', abstract)
abstract = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', abstract)
#text = re.sub(r'\[[0-9]*\]', ' ', text)
#text = re.sub(r'\s+', ' ', text)  
text = text.lower()
#text = re.sub(r'\s+', ' ', text)
#text = re.sub(r'\([^)]*\)', '', text)
#text = re.sub(r'/',' ',text)
#text = re.sub("(\d+)","",text) 
text = re.sub(r'\([^()]*\d+[^()]*\)', '', text)
text = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', text)
text = text.replace("ark.", '')


#text = text.replace("-", "") # Özel karakterleri silme
#text = " ".join(text.split()) # Gereksiz boşlukları silme


#text = re.sub(r'[\(\)]', '', text) #parantez işaretlerini silme
#text = re.sub(r'\d+', '', text) #parantez içi silme
#text = re.sub(r',', '', text) #virgül kaldır    

# Metni cümlelere ayır
cümleler = nltk.sent_tokenize(text)

# Metnin TF-IDF matrisine dönüştürülmesi

vectorizer = TfidfVectorizer(stop_words="english")
tfidf = vectorizer.fit_transform(cümleler)

# LSA modelinin oluşturulması ve uygulanması
lsa_model = TruncatedSVD(n_components=2)
lsa_matris = lsa_model.fit_transform(tfidf.toarray())

# Özetlenmiş metnin oluşturulması
özetlenmiş_cümleler = []
for i in np.argsort(lsa_matris[:, 1])[-8:]:
    özetlenmiş_cümleler.append(cümleler[i])

özetlenmiş_metin = ' '.join(özetlenmiş_cümleler)

print("Özetlenmiş metin:")
print(özetlenmiş_metin)

from rouge import Rouge
ROUGE = Rouge()

print(ROUGE.get_scores(özetlenmiş_metin, abstract))

gercekmetinuzunlugu=len(text)
özetuzunlugu=len(özetlenmiş_metin)
orjinalozet=len(abstract)


print(orjinalozet)
print(özetuzunlugu)

'''
from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/LSA.xlsx')
ws = wb['Sayfa1']
yayın1 = " Sağlık Akademisyenleri Dergisi/Covid/10.52880-sagakaderg.906994-1674600"
yayın2 = ""

recall = 0.54
precision = 0.40

ws.cell(row=57,column=3).value = yayın1
ws.cell(row=57,column=4).value = yayın2
ws.cell(row=57,column=5).value = gercekmetinuzunlugu
ws.cell(row=57,column=6).value = özetuzunlugu
ws.cell(row=57,column=8).value = orjinalozet
ws.cell(row=57,column=9).value = recall
ws.cell(row=57,column=10).value = precision

wb.save('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/LSA.xlsx')

'''

'''
ozet = özetlenmiş_metin
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Çalışmanın Özetleri/LSA/Sağlık Akademisyenleri Dergisi/Covid/10.52880-sagakaderg.906994-1674600.txt", "w",encoding="utf-8") as file:
    file.write(ozet)
  
'''