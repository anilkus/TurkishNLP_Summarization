import networkx as nx
from sklearn.feature_extraction.text import TfidfVectorizer
from nltk.tokenize import sent_tokenize
import sumy
from sumy.parsers.plaintext import PlaintextParser
from sumy.nlp.tokenizers import Tokenizer
from sumy.summarizers.lex_rank import LexRankSummarizer

import spacy
from spacy.lang.en.stop_words import STOP_WORDS
from string import punctuation
import fitz
import bs4 as bs
import urllib.request
import re
import nltk


with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Abstracts/z_Uludağ Üniversitesi Tıp Fakültesi Dergisi/979546-1911832.txt", 'r', encoding="utf-8") as text2:
    abstract = text2.read()
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/İşlenecek Kısım/z_Uludağ Üniversitesi Tıp Fakültesi Dergisi/979546-1911832.txt", 'r', encoding="utf-8") as text:
    islenecek = text.read()
    text = islenecek
    
    
abstract=abstract.lower()
abstract = re.sub(r'\([^()]*\d+[^()]*\)', '', abstract)
abstract = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', abstract)
#text = re.sub(r'\[[0-9]*\]', ' ', text)
#text = re.sub(r'\s+', ' ', text)  
text = text.lower()
#text = re.sub(r'\s+', ' ', text)
#text = re.sub(r'\([^)]*\)', '', text)
#text = re.sub(r'/',' ',text)
#text = re.sub("(\d+)","",text) 
text = re.sub(r'\([^()]*\d+[^()]*\)', '', text)
text = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', text)
text = text.replace("ark.", '')

stopwords=["Tablo","tablo","ark.","ark","ya","/","(1)"]
print(stopwords)

text = re.sub(r'\b(' + '|'.join(stopwords) + r')\b', '', text)

    
def text_summarizer(text, ratio=0.4):

    sent_list = sent_tokenize(text)
    word_list = " ".join(sent_list).split()
    word_freq = dict(nltk.FreqDist(word_list))
    G = nx.Graph()
    for word in word_freq:
        G.add_node(word, weight=word_freq[word])
    rank = nx.pagerank(G, alpha=0.85)    
    summarize_text = []
    for i in sorted(rank, key=rank.get, reverse=True):
        summarize_text.append(i)
    return " ".join(summarize_text[:int(len(summarize_text) * ratio)])

#TextRank" algoritmasını 
print(text_summarizer(text))
summary=text_summarizer(text)
from rouge import Rouge
ROUGE = Rouge()

print(ROUGE.get_scores(summary, abstract))

print("orjinal özet: ", len(abstract))
print("orjinal metin: ", len(text))
print("olusturulan özet: ", len(summary))

gercekmetinuzunlugu=len(text)
özetuzunlugu=len(summary)
orjinalozet=len(abstract)


'''
# Rouge hesaplama
rouge = Rouge()
scores = rouge.get_scores(abstract, summary)
# ROUGE-1, ROUGE-2 ve ROUGE-L skorlarına erişim
rouge_1_score = scores[0]['rouge-1']['f']
rouge_2_score = scores[0]['rouge-2']['f']
rouge_l_score = scores[0]['rouge-l']['f']


# Sonuçları yazdırma
print("ROUGE-1 Score:", rouge_1_score)
print("ROUGE-2 Score:", rouge_2_score)
print("ROUGE-L Score:", rouge_l_score)

'''

'''
from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/ASYU IEEE/Veri Ön İşleme/ÇALIŞMA ÖZETLERİ/sonuçlar/TEXTRANK.xlsx')
ws = wb['Sayfa1']
yayın1 = "Sdü Sağlık Bilimleri Dergis "
yayın2 = "Parkinson Hastalığı Patogenezinde Esansiyel Yağ Asitleri ve Kolesterolün Etkileri"

recall = 0.
precision = 0.

ws.cell(row=5,column=3).value = yayın1
ws.cell(row=5,column=4).value = yayın2
ws.cell(row=5,column=5).value = gercekmetinuzunlugu
ws.cell(row=5,column=6).value = özetuzunlugu
ws.cell(row=5,column=8).value = orjinalozet
ws.cell(row=5,column=9).value = recall
ws.cell(row=5,column=10).value = precision

wb.save('C:/Users/LENOVO/Desktop/ASYU IEEE/Veri Ön İşleme/ÇALIŞMA ÖZETLERİ/sonuçlar/TEXTRANK.xlsx')
'''

'''
ozet = summary
with open("C:/Users/LENOVO/Desktop/ASYU IEEE/Veri Ön İşleme/ÇALIŞMA ÖZETLERİ/KAYDEDİLEN ÖZETLER/TEXTRANK/SDÜ/Beyin/10.22312-sdusbed.626176-991185.txt", "w",encoding="utf-8") as file:
    file.write(ozet)

'''