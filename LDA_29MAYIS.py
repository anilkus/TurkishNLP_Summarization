import nltk
from nltk.corpus import stopwords
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.stem import WordNetLemmatizer
import bs4 as bs
import urllib.request
import re
import nltk
import networkx as nx

# Metin dosyasını okuyun ve ön işleme adımlarını uygulayın:
def preprocess_text(text,abstract):
    # Küçük harfe çevirme
    text = text.lower()
    abstract = abstract.lower()
    abstract = abstract.replace("-", "")  
    abstract = " ".join(abstract.split()) 
    
    
    # Kelime ve noktalama işaretlerine ayırma
    tokens = word_tokenize(text)
    abstract = re.sub(r'\([^()]*\d+[^()]*\)', '', abstract)
    abstract = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', abstract)
 
    text = re.sub(r'\([^()]*\d+[^()]*\)', '', text)
    text = re.sub(r'\[[^\[\]]*\d+[^\[\]]*\]', '', text)
    text = text.replace("ark.", '')
    
    # Stop kelimeleri ve noktalama işaretlerini çıkar
    stop_words = stopwords.words("turkish")
    tokens = [word for word in tokens if word.isalnum() and word not in stop_words]
    
    # Köklerine ayırma
    lemmatizer = WordNetLemmatizer()
    tokens = [lemmatizer.lemmatize(word) for word in tokens]
    
    return tokens

# Metin dosyasını oku
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Abstracts/Çukurova Medical Journal/Covid/10.17826-cumj.793835-1286453.txt", 'r', encoding="utf-8") as text2:
    abstract = text2.read()
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/İşlenecek Kısım/Çukurova Medical Journal/Covid/10.17826-cumj.793835-1286453.txt", 'r', encoding="utf-8") as text:
    islenecek = text.read()
    text = islenecek

# Metni ön işleme adımlarından geçir
preprocessed_text = preprocess_text(text,abstract)

# Belirli bir yüzdeyle cümle sayısı belirleme
percentage = 0.15 # Özetin yüzde kaçı alınacak
sentences = sent_tokenize(text)
num_sentences = int(len(sentences) * percentage)

# Cümleleri birleştirerek özet oluşturma
summary = " ".join(sentences[:num_sentences])

# Özet sonucunu yazdır
print("Metin Özeti:")
print(summary)

from rouge import Rouge
ROUGE = Rouge()

print(ROUGE.get_scores(summary, abstract))

gercekmetinuzunlugu=len(text)
özetuzunlugu=len(summary)
orjinalozet=len(abstract)


print(orjinalozet)
print(özetuzunlugu)
print("gerçek metin: ")
print(len(islenecek))



'''

from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/LDA.xlsx')
ws = wb['Sayfa1']
yayın1 = "Journal of Nursology/Covid/10.17049-ataunihem.728544-1180763"
yayın2 = ""

recall = 0.5
precision = 0.15

ws.cell(row=52,column=3).value = yayın1
ws.cell(row=52,column=4).value = yayın2
ws.cell(row=52,column=5).value = gercekmetinuzunlugu
ws.cell(row=52,column=6).value = özetuzunlugu
ws.cell(row=52,column=8).value = orjinalozet
ws.cell(row=52,column=9).value = recall
ws.cell(row=52,column=10).value = precision

wb.save('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/LDA.xlsx')
'''

'''
ozet = summary
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Çalışmanın Özetleri/LDA/Çukurova Medical Journal/Covid/10.17826-cumj.793835-1286453.txt", "w",encoding="utf-8") as file:
    file.write(ozet)

'''
